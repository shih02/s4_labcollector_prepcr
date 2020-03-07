from flask import Flask, render_template, request
from werkzeug.utils import secure_filename
import requests
import re
import H
import json
import os
import csv
from openpyxl import load_workbook, workbook

app = Flask(__name__)

@app.route('/')
def home():
    return render_template("home.html")

@app.route('/search', methods=["POST","GET"])
def result():
    if request.method == 'POST':
        samplelist = ""
        dict_list = []
        f = request.files['name']
        f.save(secure_filename(f.filename))        
        wb = load_workbook(filename = f)
        ws = wb.worksheets[0]

        for x in range(1, ws.max_row+1):
            id = ws.cell(row = x, column = 1).value
            samplelist = samplelist + id +","

        #generate response and set response variables 
        # // careful with this loop as it can break LC server!!
        url = H.url_v2 + samplelist
        response = requests.request("GET", url, headers=H.headers)
        response_text = response.text.replace("'",'"')
        my_response = json.loads(response_text)
        # //

        #rebuild response text to set count:lab id:volume
        sample = samplelist.split(",")
        for i in range(x):
            dict_list.append(dict(ID=sample[i], conc=my_response[i]['volume'], num=i+1))
            
        return render_template("result.html", dict_list=dict_list)
    else: 
        return render_template("search.html")

@app.route('/update', methods=["POST","GET"])
def update():
    if request.method == 'POST':
        samplelist = ""
        dict_list = []
        f = request.files['name']
        f.save(secure_filename(f.filename))        
        
        with open(f.filename, 'r') as csvfile:
            csvreader = csv.reader(csvfile)
            rows = list(csvreader)  
            for x in range(7,103):
                try:
                    id = rows[x][0]
                    if rows[x][8] == "FE":
                        ext = "_10"
                    elif rows[x][8] == "SE":
                        ext = "_20"
                    elif rows[x][8] == "TE":
                        ext = "_30"
                    elif rows[x][8] == "FO":
                        ext = "_40"
                    else:
                        ext = ""
                except IndexError:
                    break
                
                samplelist += id + ext + ","
        
        #generate response and set response variables 
        # // careful with this loop as it can break LC server!!
        url = H.url_v2 + samplelist
        response = requests.request("GET", url, headers=H.headers)
        response_text = response.text.replace("'",'"')
        my_response = json.loads(response_text)
        # //

        for y in range(7,103):
            try:
                id = rows[y][0]
                concentration = rows[y][4]
                if id == my_response[y-7]['label']:
                    payload = {'comments': 'payload_testing' + str(y),'origin': concentration,'volume': concentration}
                    put_response = requests.request("PUT", H.put_url+my_response[y-7]['count'], headers=H.headers, data = payload)
                    dict_list.append(dict(ID=id, conc=concentration,num=y-6, ct=my_response[y-7]['count']))
            
            except IndexError:
                break
        
        return render_template("result.html", dict_list=dict_list)
    else:
        return render_template("update.html")
        
if __name__ == "__main__":
    app.run(debug=True)