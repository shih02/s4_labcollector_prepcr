import requests
import re
import H
from openpyxl import load_workbook, workbook

#parse column A for sample IDs
samplelist = ""
wb = load_workbook(filename = 'Book1.xlsx')
ws = wb['Sheet1']
for x in range(1, ws.max_row+1):
  id = ws.cell(row = x, column = 1).value
  concentration = ws.cell(row = x, column = 2).value
  url = H.url + id
  response = requests.request("GET", url, headers=H.headers)
  res_split = re.split(r'[,:""]',response.text)
  print(res_split)
  count = res_split[4]
  payload = {'comments': 'VS testing' + str(x),'origin': concentration,'volume': concentration}
  print(id+count+str(concentration))
  put_url = "http://10.95.2.101/lab/webservice/v1/samples/" + count
  put_response = requests.request("PUT", put_url, headers=H.headers, data = payload)


# get below
# url = "http://10.95.2.101/lab/webservice/v1/samples?label=" + name
# response_text = requests.request("GET", url, headers=headers).text








