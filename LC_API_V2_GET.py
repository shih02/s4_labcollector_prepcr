import requests
import re
import H
import pprint

from openpyxl import load_workbook, workbook

#parse column A for sample IDs
samplelist = ""
wb = load_workbook(filename = 'Book1.xlsx')
ws = wb['Sheet1']
for x in range(1, ws.max_row+1):
	id = ws.cell(row = x, column = 1).value
	samplelist = samplelist + id +","

#API header for labcollector
# headers = {
#   'Accept': 'application/json;odata.metadata=full',
#   'X-LC-APP-Auth': 'b970d14af941f18aa1417c874e0f414db437fa6b277bf4c7ec196a189fc9f0c5',
#   'Host': '10.95.2.101',
#   'Accept-Encoding': 'gzip, deflate, br',
#   'Connection': 'keep-alive'
# }

sample = samplelist.split(",")
print("Loading LabCollector")

#generate response and set response variables
url = H.url_v2 + samplelist
response = requests.request("GET", url, headers=H.headers)
response_code = response.status_code
response_text = response.text
dict_list = []

#rebuild response text to set count:lab id:volume
for i in range(x):
  result = response_text.find(sample[i])
  res_split = re.split(r'[,:]',response_text[result-53:result+195])
  count = res_split[1]
  volume = res_split[21]
  sample_dict = (count+","+sample[i]+","+volume)
  dict_list.append(sample_dict)

print(response_text)
print(res_split)
print(dict_list)