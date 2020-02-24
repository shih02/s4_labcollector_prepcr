import requests
import re
import H
import json
from openpyxl import load_workbook, workbook

samplelist = ""
dict_list = []
wb = load_workbook(filename = 'Book1.xlsx')
ws = wb['Sheet1']

for x in range(1, ws.max_row+1):
    id = ws.cell(row = x, column = 1).value
    samplelist = samplelist + id +","

#generate response and set response variables 
# // careful with this loop as it can break LC server!!
url = H.url_v2 + samplelist
response = requests.request("GET", url, headers=H.headers)
response_text = response.text.replace("'",'"')
my_response = json.loads(response_text)
# //

for x in range(1, ws.max_row+1):
	id = ws.cell(row = x, column = 1).value
	concentration = ws.cell(row = x, column = 2).value
	if id == my_response[x-1]['label']:
		payload = {'comments': 'VS dictionary testing' + str(x),'origin': concentration,'volume': concentration}
		put_response = requests.request("PUT", H.put_url+my_response[x-1]['count'], headers=H.headers, data = payload)


# print(dict_list)



# get below
# url = "http://10.95.2.101/lab/webservice/v1/samples?label=" + name
# response_text = requests.request("GET", url, headers=headers).text








